
import os
import sys, time
sys.path.append("../Membersgram")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

EXCEL_PATH = 'TestResults.xlsx'

# رنگ‌ها
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

def get_valid_test_names():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    return [cell.value for cell in ws[1] if cell.value != "Date"]

def log_test_result(test_name, status):
    # if not os.path.exists(EXCEL_PATH):
    #     print("Excel file not found.")
    #     return

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    now_time = datetime.now()
    now_time_str = now_time.strftime("%H:%M:%S")
    today_str = now_time.strftime("%Y-%m-%d")

    valid_tests = get_valid_test_names()
    if test_name not in valid_tests:
        print(f"Test name '{test_name}' not found in Excel headers.")
        return

    # پیدا کردن ردیف با تاریخ امروز
    target_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == today_str:
            target_row = row
            break

    if not target_row:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=today_str)

    # پیدا کردن ستون مناسب
    col_index = valid_tests.index(test_name) + 2
    cell = ws.cell(row=target_row, column=col_index)
    entry = f"{now_time_str} - {status}"

    # فقط ثبت آخرین نتیجه (پاک کردن نتایج قبلی)
    cell.value = entry

    # تنظیم رنگ پس‌زمینه بر اساس نتیجه
    if status.lower() == "pass":
        cell.fill = GREEN_FILL
    elif status.lower() == "failed":
        cell.fill = RED_FILL

    wb.save(EXCEL_PATH)























# #######################################
# import os
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# from datetime import datetime
# from google.oauth2.service_account import Credentials
# from googleapiclient.discovery import build
# from googleapiclient.errors import HttpError

# # مسیر فایل Excel شما
# EXCEL_PATH = 'Mlk.xlsx'

# # رنگ‌ها
# GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
# RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# # مشخصات API
# SERVICE_ACCOUNT_FILE = 'C:/Users/testp/OneDrive/Desktop/Membersgram/function/credentials.json'
# SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
# SPREADSHEET_ID = '17YD6mR_RZQbmk3a5LdliJzq8KPdl9Q1S'  # شناسه شیت شما
# RANGE_NAME = 'Sheet1!A1:D10'  # محدوده مورد نظر در شیت

# def get_sheets_service():
#     creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
#     service = build('sheets', 'v4', credentials=creds)
#     return service

# def get_valid_test_names():
#     wb = load_workbook(EXCEL_PATH)
#     ws = wb.active
#     return [cell.value for cell in ws[1] if cell.value != "Date"]

# def log_test_result(test_name, status):
#     if not os.path.exists(EXCEL_PATH):
#         print("Excel file not found.")
#         return

#     wb = load_workbook(EXCEL_PATH)
#     ws = wb.active
#     now_time = datetime.now()
#     now_time_str = now_time.strftime("%H:%M:%S")
#     today_str = now_time.strftime("%Y-%m-%d")

#     valid_tests = get_valid_test_names()
#     if test_name not in valid_tests:
#         print(f"Test name '{test_name}' not found in Excel headers.")
#         return

#     # پیدا کردن ردیف با تاریخ امروز
#     target_row = None
#     for row in range(2, ws.max_row + 1):
#         if ws.cell(row=row, column=1).value == today_str:
#             target_row = row
#             break

#     if not target_row:
#         target_row = ws.max_row + 1
#         ws.cell(row=target_row, column=1, value=today_str)

#     # پیدا کردن ستون مناسب
#     col_index = valid_tests.index(test_name) + 2
#     cell = ws.cell(row=target_row, column=col_index)
#     entry = f"{now_time_str} - {status}"

#     # فقط ثبت آخرین نتیجه (پاک کردن نتایج قبلی)
#     cell.value = entry

#     # تنظیم رنگ پس‌زمینه بر اساس نتیجه
#     if status.lower() == "pass":
#         cell.fill = GREEN_FILL
#     elif status.lower() == "failed":
#         cell.fill = RED_FILL

#     wb.save(EXCEL_PATH)

#     # ارسال نتایج به Google Sheets
#     try:
#         service = get_sheets_service()

#         # آماده‌سازی داده‌ها برای ارسال
#         values = [
#             [today_str, test_name, entry]  # ارسال تاریخ، نام تست و نتیجه
#         ]

#         # نوشتن به Google Sheets
#         body = {
#             'values': values
#         }
#         result = service.spreadsheets().values().append(
#             spreadsheetId=SPREADSHEET_ID,
#             range=RANGE_NAME,
#             valueInputOption='RAW',
#             body=body
#         ).execute()

#         print(f"Updated {result.get('updates').get('updatedCells')} cells in Google Sheets.")

#     except HttpError as err:
#         print(f"Error: {err}")




